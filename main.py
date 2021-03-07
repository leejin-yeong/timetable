import openpyxl 

class member:
    current_time = 0 # <= 14
    total_num = 0
    col_time = 0 # <= 8

    def __init__(self, total_num):
        self.total_num = total_num
    
    def reset_col_time(self):
        self.col_time=0

#근무자 리스트 생성
MemberList = {
    "재현": member(21), "병국": member(16), "예윤": member(20),
    "혜지": member(23), "태훈": member(19), "유진": member(20),
    "서연": member(19), "한솔": member(20), "희지": member(18),
    "현빈": member(13), "준범": member(5)
}

# total_num이 작은 순서대로 list 정렬
def SortByNum(peoplelist):
    length = len(peoplelist)
    for i in range(length):
        total = MemberList[peoplelist[i]].total_num
        for j in range(i+1, length):
            next_total = MemberList[peoplelist[j]].total_num
            if total > next_total:
                temp = peoplelist[j]
                peoplelist[j] = peoplelist[i]
                peoplelist[i] = temp
    return peoplelist

def UpdateTime(person, rows):
    if rows == 7 : 
        add = 4 # 저녁
    else: 
        add = 1.5 # 낮
    MemberList[person].current_time += add
    MemberList[person].col_time += add

# 엑셀 파일있는 경로
path = "C:/Users/cooki/Desktop/timetable/example.xlsx"

# workbook 객체
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Sheet1']
#workbook.active로 하면 현재 활성화된 시트를 뜻하는 듯.
write_wb = openpyxl.Workbook()
result = write_wb.active

people_list = []
result_list = []
cols = 1
for c in ws.columns:#같은 열부터 읽음
    rows = 1
    #모든 근무자의 하루 근무량 초기화
    for m in MemberList:
        MemberList[m].reset_col_time()

    for r in c:
        prev_list = [] #이전 타임 근무자 저장
        people = r.value.split() #공백 기준 나누기
        
        if len(people) < 3:
            result.cell(rows,cols,r.value)
            #현재 근무시간 업데이트
            for person in people:
                UpdateTime(person, rows)
        else:
            first_list = []
            second_list = []
            third_list = []
            if rows != 7:
                next_cell = ws.cell(rows+1, cols).value.split()

            #1. 이전 타임에 근무하는지 확인
            for person in people:   
                if person in prev_list: #이전 타임에 근무함 
                    if (rows != 7) and (person in next_cell): #다음 타임 근무 가능
                        first_list.insert(0, person)
                    else:
                        first_list.append(person)
                elif (rows!=7) and (person in next_cell): #다음 타임 근무 가능
                    second_list.append(person)
                else:
                    third_list.append(person)

            first_list = SortByNum(first_list)
            second_list = SortByNum(second_list)
            third_list = SortByNum(third_list)
            people_list = first_list + second_list + third_list


            for person in people_list:
                if rows == 7: 
                    add = 4
                else: 
                    add = 1.5
        
                if (MemberList[person].col_time + add <= 8) and (MemberList[person].current_time + add <= 14): 
                    #result_list.remove(person) # 인덱스가 당겨지니까 당겨진 것에 대해선 체크를 안함
                    result_list.append(person)

            result.cell(rows, cols, ' '.join(result_list[0:2]))
            #print(cols, rows, result_list[0], result_list[1])
            for i in result_list:
                UpdateTime(i, rows)
                #결과 값을 prev_list에 저장. 다음 행 우선순위 결정 시 사용
                prev_list.append(i)
            result_list = []
            people_list = []
                     
        rows += 1
        if rows > 7 : break
    cols += 1
    if cols > 5 : break
write_wb.save("reuslt.xlsx")